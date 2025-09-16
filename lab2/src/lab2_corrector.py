import os
import subprocess
import glob
import shutil
import traceback
import json
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from corrector_agent import CorrectorAgent
from criterios_correcao import get_correction_criteria, get_correction_instructions, get_main_prompt, get_refined_prompt


class CorrectionFailed(Exception):
    def __init__(self, corrector, error_type, message):
        corrector.error_type = error_type
        corrector.student_logs.append(message)
        super().__init__(message)

class CompilationError(CorrectionFailed):
    def __init__(self, corrector, message):
        super().__init__(corrector, "ERRO-COMPILACAO", message)

class RuntimeError(CorrectionFailed):
    def __init__(self, corrector, message):
        super().__init__(corrector, "ERRO-RUNTIME", message)

class WrongFilePathError(CorrectionFailed):
    def __init__(self, corrector, message):
        super().__init__(corrector, "ARQUIVO-NOME-ERRADO", message)

class OutputFormattingError(CorrectionFailed):
    def __init__(self, corrector, message):
        super().__init__(corrector, "FORMATACAO-OUTPUT-ERRADA", message)

class FailedTestcaseError(CorrectionFailed):
    def __init__(self, corrector, message):
        super().__init__(corrector, "ERRO-CASOS-TESTE", message)

class Lab2Corrector():
    def __init__(self, alunos_path, testcases_path, numero_lab, use_ai, aluno=None):
        self.alunos_path = alunos_path
        self.remove_error_type_txts()
        self.create_student_folders()
        if aluno is None:
            self.alunos_list = sorted(os.listdir(self.alunos_path))
        else:
            self.alunos_list = [aluno]
        self.testcases_path = testcases_path
        self.numero_lab = numero_lab
        self.use_ai = use_ai
        self.aluno = aluno
        self.error_type = "No-Errors"
        self.student_logs = []

        self.wb = Workbook()
        self.ws = self.wb.active

        self.headers = [
            "", "Nota final", "Prazo", "Arquivo", "Saída", "Identação", "Bronco",
            "Índice 0", "Str. Descrição", "Agenda", "R.A.TAD", "L/E/main",
            "Global", "Busca Binária", "Func. Públicas", "fclose/free",
            "Outros", "Observações (sem desconto na nota)"
        ]

    def print_logs(self, aluno_path, tipo_correcao, encoding):
        logs_correcao_path = aluno_path + f"/logs_correcao_{tipo_correcao}.txt"
        with open(logs_correcao_path, "w", encoding=encoding) as logs:
            print(f"{self.error_type}\n\n", file=logs)
            for error in self.student_logs:
                print(error, file=logs)

    def add_error_message(self, message):
        self.student_logs.append(message)

    def create_student_folders(self):
        lowercase_files = glob.glob(os.path.join(self.alunos_path, "lab*"))
        for file_path in lowercase_files:
            filename = os.path.basename(file_path)
            new_filename = filename.replace('lab', 'Lab')
            os.rename(file_path, os.path.join(self.alunos_path, new_filename))

        CPP_files = glob.glob(os.path.join(self.alunos_path, "Lab*.CPP"))
        for file_path in CPP_files:
            filename = os.path.basename(file_path)
            new_filename = filename.replace('.CPP', '.cpp')
            os.rename(file_path, os.path.join(self.alunos_path, new_filename))

        cpp_files = glob.glob(os.path.join(self.alunos_path, "Lab*.cpp"))
        for file_path in cpp_files:
            filename = os.path.basename(file_path) 
            folder_name = filename[5:].replace('.cpp','')
            new_folder_path = os.path.join(self.alunos_path, folder_name)
            os.makedirs(new_folder_path, exist_ok=True)
            shutil.move(file_path, os.path.join(new_folder_path, filename))  

    def remove_outputs_folder(self, aluno_path):
        outputs_path = os.path.join(aluno_path, "outputs")
        if os.path.exists(outputs_path) and os.path.isdir(outputs_path):
            shutil.rmtree(outputs_path)

    def remove_unwanted_files(self, aluno_path):
        for f in os.listdir(aluno_path):
            full_path = os.path.join(aluno_path, f)
            if f == "a.out":
                os.remove(full_path)
            elif f.endswith(".txt") and not f.startswith("logs_correcao"):
                os.remove(full_path)
                
    def remove_error_type_txts(self):
        for filename in os.listdir(self.alunos_path):
            file_path = os.path.join(self.alunos_path, filename)
            if os.path.isfile(file_path) and filename.endswith(".txt"):
                os.remove(file_path)

    def add_to_error_type_txt(self, aluno_path):
        error_file_path = os.path.join(self.alunos_path, f"{self.error_type}.txt")
        with open(error_file_path, "a", encoding="utf-8") as f:
            f.write(os.path.basename(aluno_path) + "\n")

    def get_and_handle_output(self, aluno_path, testcase):
        output_path = glob.glob(f'{aluno_path}/Lab*.txt')
        if not output_path:
            raise FailedTestcaseError(self, "Nao criou o arquivo txt de saida\n")
        output_folder = os.path.join(aluno_path, 'outputs')
        os.makedirs(output_folder, exist_ok=True)
        final_output_path = os.path.join(output_folder, f"{testcase}.txt")
        shutil.copy(output_path[0], final_output_path)
        os.remove(output_path[0])
        with open(final_output_path, encoding='latin-1') as output_file:
            return output_file.readlines()
        
    def get_student_code(self, aluno_path):
        cpp_path = glob.glob(f'{aluno_path}/Lab*.cpp')
        if not cpp_path:
            raise WrongFilePathError(self, "Nao enviou o arquivo .cpp\n")
        with open(cpp_path[0], encoding='latin-1') as cpp_file:
            return cpp_file.read()

    def compile_student_code(self, aluno_path):
        if os.path.isdir(aluno_path):
            cpp_path = os.path.join(f'{aluno_path}/Lab*.cpp')
            try:
                result = subprocess.run(
                    f'gcc {cpp_path} -o {aluno_path}/a.out',
                    shell=True, 
                    check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE,
                    timeout=5
                )
                if result.returncode == 0 and result.stderr:
                    self.add_error_message(f"WARNINGS NA COMPILACAO:\n{result.stderr.decode('utf-8', errors='replace')}")
            except subprocess.CalledProcessError as e:
                raise CompilationError(self, f"Codigo de saida: {e.returncode}\n")
            except subprocess.TimeoutExpired:
                raise CompilationError(self, f"Tempo limite de compilacao excedido.", aluno_path)
            except Exception as e:
                raise CompilationError(self, f"Ocorreu um erro inesperado na compilacao: {e}\n")

    def run_student_code(self, aluno_path, testcase):
        try:
            testcase_path = os.path.join(self.testcases_path, testcase)
            subprocess.run(
                f'cp {testcase_path}/entrada* {aluno_path}',
                shell=True, 
                check=True
            )
            subprocess.run(
                './a.out', 
                cwd=aluno_path, 
                shell=True, 
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=5
            )
        except subprocess.CalledProcessError as e:
            raise RuntimeError(self, f"Erro na execucao do caso teste {testcase}.\nCodigo de saida: {e.returncode}\n")
        except subprocess.TimeoutExpired:
            raise RuntimeError(self, f"Tempo limite de execucao do caso teste {testcase} excedido")
        except Exception as e:
            raise RuntimeError(self, f"Ocorreu um erro inesperado na execucao do caso teste {testcase}\n")

    def check_fopen_path(self, student_code):
        pattern_entrada = fr'fopen\s*\(\s*"entrada{self.numero_lab}\.txt"\s*,\s*".*?"\s*\)'
        pattern_saida = fr'fopen\s*\(\s*"Lab{self.numero_lab}_[a-zA-Z0-9_]+\.txt"\s*,\s*".*?"\s*\)'
        nome_entrada_correto = re.search(pattern_entrada, student_code) is not None
        nome_saida_correto = re.search(pattern_saida, student_code) is not None
        if not nome_entrada_correto or not nome_saida_correto:
            raise WrongFilePathError(self, "Erro no nome dos arquivos de entrada ou saída\n")

    def correct_code(self, aluno_path, aluno_row):
        code = self.get_student_code(aluno_path)
        self.check_fopen_path(code)
        if self.use_ai:
            response = self.do_ai_correction(aluno_path, code)
            json_str = "\n".join(response)
            response_dict = json.loads(json_str)
        
    def correct_output(self, aluno_path, testcase):
        self.run_student_code(aluno_path, testcase)
        output = self.get_and_handle_output(aluno_path, testcase)
        self.test_formatacao(testcase, output)
        self.compare_with_testcase(testcase, output)

    def test_formatacao(self, testcase, output):
        linhas = [linha.rstrip('\n') for linha in output] 
        pos = 0

        if len(linhas) < 4:
            raise OutputFormattingError(self, f"Esperadas 4 linhas de cabeçalho no caso teste {testcase}")
        pos += 4

        if pos >= len(linhas) or [w.upper() for w in linhas[pos].split()] != ["FLIGHT", "FROM"]:
            print([w.upper() for w in linhas[pos].split()])
            raise OutputFormattingError(self, f"Linha 'FLIGHT FROM' ausente no caso teste {testcase}")
        pos += 1

        if pos >= len(linhas) or linhas[pos].strip() != "":
            raise OutputFormattingError(self, f"Faltando linha em branco após 'FLIGHT FROM' no caso teste {testcase}")
        pos += 1

        while pos < len(linhas) and linhas[pos].strip() != "":
            campos = linhas[pos].split()
            if len(campos) < 2 or not campos[0].isdigit():
                raise OutputFormattingError(self, f"Bloco de voos autorizados inválido no caso teste {testcase}")
            pos += 1

        if pos >= len(linhas) or linhas[pos].strip() != "":
            raise OutputFormattingError(self, f"Faltando linha em branco após bloco de voos autorizados no caso teste {testcase}")
        pos += 1

        if pos >= len(linhas) or [w.upper() for w in linhas[pos].split()] != ["SITUACAO", "DA", "FILA"]:
            raise OutputFormattingError(self, f"Linha 'Situacao da fila' ausente no caso teste {testcase}")
        pos += 1

        if pos >= len(linhas) or linhas[pos].strip() != "":
            raise OutputFormattingError(self, f"Faltando linha em branco após 'Situacao da fila' no caso teste {testcase}")
        pos += 1

    def compare_with_testcase(self, testcase, output):
        linhas = [linha.rstrip('\n') for linha in output] 

        student_authorized_flights = []
        student_pending_flights = []
        student_flight_origins = {}
        
        reading_authorized_flights = False
        reading_pending_flights = False

        i = 0
        while i < len(linhas):
            splitted_line = linhas[i].split()
            if linhas[i].strip() == "":
                i += 1
                continue
            if [w.upper() for w in splitted_line] == ["FLIGHT", "FROM"]:
                reading_authorized_flights = True
                i += 1
                continue
            elif [w.upper() for w in splitted_line] == ["SITUACAO", "DA", "FILA"]:
                reading_authorized_flights = False
                reading_pending_flights = True
                i += 1
                continue
            if reading_authorized_flights:
                student_authorized_flights.append(splitted_line[0])
                if splitted_line[0] != "0000":
                    student_flight_origins[splitted_line[0]] = " ".join(splitted_line[1:])
            elif reading_pending_flights:
                student_pending_flights.append(splitted_line[0])
                if splitted_line[0] != "0000":
                    student_flight_origins[splitted_line[0]] = " ".join(splitted_line[1:])
            i += 1

        answers_path = os.path.join(self.testcases_path, testcase, 'saida2.json')
        with open(answers_path, "r", encoding="utf-8") as answers_file:
            answers = json.load(answers_file)
        if student_authorized_flights != answers["ordem_voos"]["authorized"]:
            raise FailedTestcaseError(self, f"Falhou no caso teste {testcase}: ordem das viagens AUTORIZADAS errada")
        if answers["ordem_voos"]["pending"]:
            if student_pending_flights != answers["ordem_voos"]["pending"]:
                raise FailedTestcaseError(self, f"Falhou no caso teste {testcase}: ordem das viagens PENDENTES errada")
            if student_flight_origins != answers["flight_origins"]:
                raise FailedTestcaseError(self, f"Falhou no caso teste {testcase}: DESTINO das viagens está errado")
        else:
            if not student_pending_flights:
                raise FailedTestcaseError(self, f"Falhou no caso teste {testcase}: sem mensagem de PENDENTES VAZIA")
            student_filtered = {k: v for k, v in student_flight_origins.items()
                    if str(k).isdigit() and len(str(k)) == 4}
            if student_filtered != answers["flight_origins"]:
                raise FailedTestcaseError(self, f"Falhou no caso teste {testcase}: DESTINO das viagens está errado")

    def do_ai_correction(self, aluno_path, student_code):
        correction_criteria_prompt = get_correction_criteria()
        correction_instructions_prompt = get_correction_instructions()
        corrector_agent = CorrectorAgent(aluno_path)
        prompt = get_main_prompt(correction_criteria_prompt, correction_instructions_prompt, student_code)
        response = corrector_agent.respond(prompt)
        refined_prompt = get_refined_prompt(correction_criteria_prompt, correction_instructions_prompt, student_code, response)
        refined_response = corrector_agent.respond(refined_prompt)
        for line in response:
            self.print_log(line, aluno_path, tipo_correcao="ia", encoding='utf-8')
        self.print_log('', aluno_path, tipo_correcao="ia", encoding='utf-8')
        for line in refined_response:
            self.print_log(line, aluno_path, tipo_correcao="ia", encoding='utf-8')
        self.print_log('', aluno_path, tipo_correcao="ia", encoding='utf-8')
        return refined_response

    def make_student_correction(self, aluno_path, progress):
        self.error_type = "No-Errors"
        self.student_logs = []
        aluno_row = progress + 1
        try:
            self.correct_code(aluno_path, aluno_row)
            self.compile_student_code(aluno_path)
        except (CompilationError, WrongFilePathError):
            return
        for testcase in os.listdir(self.testcases_path):
            try:
                self.correct_output(aluno_path, testcase)
            except (FailedTestcaseError, RuntimeError, OutputFormattingError):
                continue

    def make_correction(self):
        progress = 1
        try:
            for aluno in self.alunos_list:
                print(f"Correcting... ({progress}/{len(self.alunos_list)}). Current student: {aluno}")
                aluno_path = os.path.join(self.alunos_path, aluno)
                progress += 1
                self.remove_outputs_folder(aluno_path)
                self.make_student_correction(aluno_path, progress)
                self.remove_unwanted_files(aluno_path)
                self.add_to_error_type_txt(aluno_path)
                self.print_logs(aluno_path, tipo_correcao="auto", encoding="utf-8")
            print("Correction ended successfully")
        except Exception as e:
            print(f"Correction failed due to error: {e}")
            traceback.print_exc()



